import os
import requests
import pandas as pd
from datetime import datetime, timedelta, timezone
from bs4 import BeautifulSoup
from dotenv import load_dotenv
import time
from email.utils import parsedate_to_datetime  # 시스템 로케일 영향을 받지 않는 날짜 분석 도구
from openpyxl.styles import Alignment, PatternFill

# .env 파일에서 환경 변수를 로드합니다.
try:
    load_dotenv()
    CLIENT_ID = os.environ["NAVER_CLIENT_ID"]
    CLIENT_SECRET = os.environ["NAVER_CLIENT_SECRET"]
except KeyError:
    CLIENT_ID = None
    CLIENT_SECRET = None
    print("Warning: Naver API credentials not found. .env 파일을 정확히 설정했는지 확인해주세요.")

KEYWORDS = ["신세계면세점", "롯데면세점", "신라면세점", "현대면세점", "인천공항공사", "한국관광공사", "한국면세점협회", "패션", "K뷰티"]

# 간단한 감성 분석을 위한 단어 리스트
POS_WORDS = ['상승', '증가', '성장', '확대', '최대', '호조', '기대', '인기', '강화', '돌파', '활성화', '유치', '맞춤', '프로모션', '혜택']
NEG_WORDS = ['하락', '감소', '위기', '악화', '부진', '침체', '우려', '논란', '적자', '축소', '불만', '피해']

def analyze_sentiment(text):
    """간단한 키워드 기반으로 텍스트의 감성을 분석합니다."""
    if not text:
        return '중립'
    pos_count = sum(1 for word in POS_WORDS if word in text)
    neg_count = sum(1 for word in NEG_WORDS if word in text)
    if pos_count > neg_count:
        return '긍정'
    elif neg_count > pos_count:
        return '부정'
    else:
        return '중립'

def fetch_naver_news(keyword, target_date):
    """
    특정 키워드와 날짜(어제)를 기준으로 네이버 뉴스를 수집합니다.
    - 시간대(Timezone) 문제를 해결하여 KST 기준 날짜로 정확히 필터링합니다.
    - API 정렬 불완전성 문제를 보완하여 조기 종료를 방지합니다.
    """
    if not CLIENT_ID or not CLIENT_SECRET:
        print("Error: Naver API credentials are not configured.")
        return []

    url = "https://openapi.naver.com/v1/search/news.json"
    headers = {
        "X-Naver-Client-Id": CLIENT_ID,
        "X-Naver-Client-Secret": CLIENT_SECRET
    }

    all_articles = []
    start = 1
    display = 100
    print(f"Fetching news for: '{keyword}' on {target_date.strftime('%Y-%m-%d')}")

    # 한국 표준시(KST) 보정을 위한 설정
    target_date_only = target_date.date()
    # API 응답은 최대 1000개까지만 가능
    while start <= 1000:
        params = {"query": keyword, "display": display, "start": start, "sort": "date"}
        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()
        except Exception as e:
            print(f"Error fetching data: {e}")
            break

        items = data.get('items', [])
        if not items:
            break

        has_reached_far_past = False
        for item in items:
            try:
                # 1. UTC 기준 시간을 파싱한 후 한국 시간(KST, UTC+9)으로 변환합니다.
                pub_date_utc = parsedate_to_datetime(item['pubDate'])
                # Naive datetime 객체일 경우 UTC로 간주하고 KST로 변환
                if pub_date_utc.tzinfo is None:
                     pub_date_kst = pub_date_utc.replace(tzinfo=timezone.utc).astimezone(tz=None)
                else:
                     pub_date_kst = pub_date_utc.astimezone(None) # 시스템 로컬 시간대(한국)로 변환
            except Exception:
                # 날짜 파싱에 실패하면 해당 기사는 건너뜁니다.
                continue

            pub_date_only = pub_date_kst.date()

            # 2. 목표 날짜(어제)에 작성된 기사만 필터링합니다.
            if pub_date_only == target_date_only:
                title = item['title'].replace('<b>', '').replace('</b>', '').replace('&quot;', '"')
                description = item['description'].replace('<b>', '').replace('</b>', '').replace('&quot;', '"')
                link = item['link']

                reporter, media, img_url, content = "알수없음", "알수없음", "", description

                if "n.news.naver.com" in link:
                    try:
                        news_resp = requests.get(link, headers={'User-Agent': 'Mozilla/5.0'}, timeout=5)
                        if news_resp.status_code == 200:
                            soup = BeautifulSoup(news_resp.text, 'html.parser')
                            
                            # 매체명 추출 로직 강화
                            meta_media = soup.select_one('meta[property="me2:category1"]')
                            if meta_media and meta_media.has_attr('content'):
                                media = meta_media['content']
                            else:
                                media_tag = soup.select_one('.media_end_head_top_logo img')
                                if media_tag:
                                    media = media_tag.get('title') or media_tag.get('alt') or "알수없음"
                            
                            reporter_tag = soup.select_one('.media_end_head_journalist_name')
                            if reporter_tag: reporter = reporter_tag.text.strip().split()[0]
                            
                            img_tag = soup.select_one('#img1')
                            if img_tag and img_tag.has_attr('data-src'): img_url = img_tag['data-src']
                            
                            content_tag = soup.select_one('#dic_area')
                            if content_tag: content = content_tag.text.strip()
                    except Exception:
                        pass # 상세 정보 수집에 실패해도 기본 정보는 저장

                sentiment = analyze_sentiment(title + " " + content)
                all_articles.append({
                    "헤드라인": title, "매체": media, "기자명": reporter,
                    "내용요약": content[:200] + "..." if len(content) > 200 else content,
                    "기사 링크": link, "기사 사진 링크": img_url,
                    "감성": sentiment, "본문": content,
                    "작성일": pub_date_kst.strftime("%Y-%m-%d %H:%M:%S")
                })
                
                # 키워드당 최대 30개 기사 수집 제한
                if len(all_articles) >= 30:
                    break
            
            # 3. 중요: 조기 종료 조건 완화
            # API 정렬이 완벽하지 않으므로, 목표일보다 한참 전(예: 2일 전) 기사가 나오면 종료
            elif pub_date_only < target_date_only - timedelta(days=2):
                has_reached_far_past = True
                break
        
        # 키워드당 30개를 채웠거나, 아주 오래된 기사를 만나면 다음 키워드로 넘어감
        if len(all_articles) >= 30 or has_reached_far_past:
            break

        start += display
        time.sleep(0.1) # API 호출 간격

    return all_articles


def main():
    """스크립트의 메인 실행 함수. 키워드별로 뉴스를 수집하여 엑셀 파일로 저장합니다."""
    print("Starting Data Collection...")
    yesterday = datetime.now() - timedelta(days=1)
    target_date_str = yesterday.strftime("%Y-%m-%d")
    file_name = f"{target_date_str}_뉴스모니터링.xlsx"

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        for keyword in KEYWORDS:
            articles = fetch_naver_news(keyword, yesterday)
            df = pd.DataFrame(articles)

            if df.empty:
                # 데이터가 없을 경우 빈 프레임과 컬럼만 생성
                df = pd.DataFrame(columns=["헤드라인", "매체", "기자명", "내용요약", "기사 링크", "기사 사진 링크", "감성", "본문", "작성일"])
            
            df.to_excel(writer, sheet_name=keyword, index=False)
            
            # 엑셀 열 너비 및 헤더 스타일 설정
            worksheet = writer.sheets[keyword]
            
            # 헤더 배경색
            header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for cell in worksheet[1]:
                cell.fill = header_fill
            
            # 열 너비 지정
            column_widths = {'A': 40, 'B': 15, 'C': 12, 'D': 50, 'E': 20, 'F': 20, 'G': 10, 'H': 50, 'I': 20}
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # 하이퍼링크 적용
            for row in range(2, len(df) + 2):
                # 기사 링크 (E열)
                link_cell = worksheet.cell(row=row, column=5)
                if link_cell.value:
                    link_cell.hyperlink = link_cell.value
                    link_cell.style = "Hyperlink"
                # 기사 사진 링크 (F열)
                img_cell = worksheet.cell(row=row, column=6)
                if img_cell.value:
                    img_cell.hyperlink = img_cell.value
                    img_cell.style = "Hyperlink"

            # 모든 셀 자동 줄바꿈 및 세로 정렬
            for row_cells in worksheet.iter_rows():
                for cell in row_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            
            print(f"  -> Sheet '{keyword}' saved with {len(df)} articles.")
    
    print(f"\nData collection complete! Saved to {file_name}")

if __name__ == "__main__":
    main()

